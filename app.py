from flask import Flask, render_template, request, send_file
import os
import pandas as pd
import uuid
import re

import zipfile
from io import BytesIO

from openpyxl import load_workbook 
from openpyxl.styles import Font, PatternFill, Alignment 
from openpyxl.utils import get_column_letter 

app = Flask(__name__)

# Configure upload and processed folders
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

# Ensure the folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

def sanitize_filename(name):
    """Sanitize user-provided filename."""
    cleaned = re.sub(r'[^a-zA-Z0-9_\-\s]', '', name.strip())
    cleaned = re.sub(r'\s+', '_', cleaned)
    return cleaned[:50] or 'processed_data'

# ✅ Add this utility function to clean numeric columns
def clean_numeric(column):
    return pd.to_numeric(
        column.astype(str)
        .str.replace(',', '', regex=False)
        .str.replace(r'[^\d\.\-]', '', regex=True),
        errors='coerce'
    ).fillna(0)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files.get('file_process')  
        custom_name = request.form.get('filename', 'processed_data')

        if not file or file.filename == '':
            return "No file uploaded", 400

        if not file.filename.endswith('.xlsx'):
            return "Only XLSX files are allowed", 400

        safe_name = sanitize_filename(custom_name)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        try:
            merged_df = process_file(file_path)
        except Exception as e:
            return f"Error processing file: {str(e)}", 500

        unique_id = uuid.uuid4().hex[:8]
        output_filename = f"temp_{unique_id}.xlsx"
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)
        merged_df.to_excel(output_path, index=False)

        return send_file(output_path, as_attachment=True, download_name=f"{safe_name}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('index.html')

def process_file(input_path):
    data = pd.read_excel(input_path)

    required_columns = {'DATE_OF_TRAN', 'RPTCODE', 'TRAN_AMT'}
    if not required_columns.issubset(data.columns):
        raise ValueError(f"Missing columns: {required_columns - set(data.columns)}")

    data['DATE_OF_TRAN'] = pd.to_datetime(data['DATE_OF_TRAN'])
    categories = data['RPTCODE'].unique()

    final_dfs = []
    for category in categories:
        filtered_data = data[data['RPTCODE'] == category]
        total_amounts = filtered_data.groupby('DATE_OF_TRAN', as_index=False)['TRAN_AMT'].sum()

        for col in ['SR', 'TRANID', 'ACCOUNT NUMBER', 'DR/CR', 'RPTCODE', 'ACCOUNT NAME']:
            total_amounts[col] = None

        total_amounts['Total'] = 'Total'
        filtered_data['Total'] = ''
        df_final = pd.concat([filtered_data, total_amounts]).sort_values(by=['DATE_OF_TRAN', 'Total'])
        df_final = df_final.drop(columns=['Total']).reset_index(drop=True)

        final_dfs.append(df_final)

    return pd.concat(final_dfs, ignore_index=True)

@app.route('/compare', methods=['POST'])
def compare_files():
    file1 = request.files.get('file1')
    file2 = request.files.get('file2')

    if not (file1 and file2):
        return "Both files are required", 400

    filepath1 = os.path.join(app.config['UPLOAD_FOLDER'], file1.filename)
    filepath2 = os.path.join(app.config['UPLOAD_FOLDER'], file2.filename)
    file1.save(filepath1)
    file2.save(filepath2)

    output_folder = app.config['PROCESSED_FOLDER']
    output_files = process_comparison(filepath1, filepath2, output_folder)

    # ✅ Create ZIP file
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
        zipf.write(output_files["common"], arcname="common.xlsx")
        zipf.write(output_files["only_in_soc"], arcname="only_in_soc.xlsx")
        zipf.write(output_files["only_in_bk"], arcname="only_in_bk.xlsx")

    zip_buffer.seek(0)
    return send_file(zip_buffer, as_attachment=True, download_name="comparison_results.zip", mimetype="application/zip")

def process_comparison(filepath1, filepath2, output_folder):
    d1 = pd.read_excel(filepath1).sort_values(by='BkAcc').reset_index(drop=True)
    d1['Slno'] = range(1, len(d1) + 1)

    d2 = pd.read_excel(filepath2).sort_values(by='BkAcc').reset_index(drop=True)
    d2['Slno'] = range(1, len(d2) + 1)

    d1['SocBal'] = d1['SocBal'].astype(str)
    d2['BkBal'] = d2['BkBal'].astype(str)

    d_common = d1.merge(d2, on="BkAcc")
    d_not_common = d1.merge(d2, on="BkAcc", how="outer", indicator=True)
    
    only_in_d1 = d_not_common[d_not_common['_merge'] == 'left_only'].drop(columns=['_merge'])
    only_in_d2 = d_not_common[d_not_common['_merge'] == 'right_only'].drop(columns=['_merge'])

    # ✅ Use clean_numeric to fix invalid values
    d_common['ScoBal'] = clean_numeric(d_common['SocBal'])
    d_common['BkBal'] = clean_numeric(d_common['BkBal'])

    d_common["Difference"] = d_common["ScoBal"] + d_common["BkBal"]
    d_common = d_common[d_common['Difference'] != 0]

    common_path = os.path.join(output_folder, "common.xlsx")
    only_d1_path = os.path.join(output_folder, "only_in_d1.xlsx")
    only_d2_path = os.path.join(output_folder, "only_in_d2.xlsx")

    d_common.to_excel(common_path, index=False)
    only_in_d1.to_excel(only_d1_path, index=False)
    only_in_d2.to_excel(only_d2_path, index=False)

    format_excel(common_path, d_common)

    return {"common": common_path, "only_in_soc": only_d1_path, "only_in_bk": only_d2_path}

def format_excel(filepath, df):
    wb = load_workbook(filepath)
    ws = wb.active

    orange_fill = PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")
    blue_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
    green_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")

    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws[f"{get_column_letter(col_idx)}1"]
        cell.font = Font(bold=True, color="000000", size=14)
        cell.alignment = Alignment(horizontal="center")

        if 1 <= col_idx <= 5:
            cell.fill = orange_fill
        elif 6 <= col_idx <= 8:
            cell.fill = blue_fill
        elif col_idx == 9:
            cell.fill = green_fill

        max_length = max(len(str(val)) for val in df[column]) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length

    wb.save(filepath)

if __name__ == '__main__':
    app.run(debug=True)
